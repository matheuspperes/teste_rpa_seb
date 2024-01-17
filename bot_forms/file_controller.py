from openpyxl import Workbook, load_workbook
from pathlib import Path
import traceback


base_path = Path(__file__).parent

def read_file():
    try:
        file_path = Path(base_path, 'modelo_bot_forms.xlsx')
        wb = load_workbook(file_path)
        ws = wb.active
        rows = ws.max_row
        
        items = []
        for row in range(2, rows + 1):
            job = ws[f'A{row}'].value
            city = ws[f'B{row}'].value
            is_effective = ws[f'C{row}'].value
            
            if is_effective not in ("Sim", "Não"): continue
            
            if job is None or city is None and is_effective is None:
                continue
            else:
                items.append((job, city, is_effective))
                
        return {'error': False, 'type': '', 'data': items}
    
    except:
        error = traceback.format_exc()
        return {'error': True, 'type': 'Erro ao ler arquivo', 'data': error}
    

def write_file(items, import_id):
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(['CARGO', 'LOCALIDADE', 'EFETIVO', 'STATUS'])
        
        for row in range(2, len(items) + 2):
            index = row - 2
            data = items[index]
            
            ws[f'A{row}'] = data[0]
            ws[f'B{row}'] = data[1]
            ws[f'C{row}'] = data[2]
            ws[f'D{row}'] = data[3]
        
        file_path = Path(base_path, 'documents', f'{import_id}.xlsx')
        
        # Validação se a pasta existe
        documents_folder = file_path.parent
        if not documents_folder.is_dir():
            documents_folder.mkdir(parents=True, exist_ok=True)
            
        wb.save(file_path)
        
        return {'error': False, 'type': '', 'data': ''}
        
    except:
        error = traceback.format_exc()
        return {'error': True, 'type': 'Erro ao escrever arquivo', 'data': error}