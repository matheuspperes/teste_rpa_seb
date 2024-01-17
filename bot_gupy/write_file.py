from openpyxl import Workbook
from pathlib import Path
import traceback


base_path = Path(__file__).parent

def write_file(items, import_id):
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(['CARGO', 'LOCALIDADE', 'EFETIVIDADE'])
        
        for row in range(2, len(items) + 2):
            index = row - 2
            data = items[index]
            
            ws[f'A{row}'] = data[0]
            ws[f'B{row}'] = data[1]
            ws[f'C{row}'] = data[2]
        
        file_path = Path(base_path, 'documents', f'{import_id}.xlsx')
        
        # Validação se a pasta existe
        documents_folder = file_path.parent
        if not documents_folder.is_dir():
            documents_folder.mkdir(parents=True, exist_ok=True)
            
        wb.save(file_path)
        
        return {'error': False, 'type': '', 'data': ''}
        
    except:
        error = traceback.format_exc()
        return {'error': True, 'type': 'Erro ao escrever arquivo', 'data': error}